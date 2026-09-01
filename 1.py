from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.table import Table, TableStyleInfo

path = "/mnt/data/SQL_Data_Engineer_Roadmap.xlsx"

levels = [
    ("LEVEL 1", "SQL Foundation", "Nền tảng truy vấn và thao tác dữ liệu", "SELECT, WHERE, DISTINCT, ORDER BY, LIMIT, INSERT, UPDATE, DELETE, NULL, COALESCE, CASE, GROUP BY, HAVING, COUNT/SUM/AVG/MIN/MAX", "Lọc đơn hàng, tính KPI, cập nhật dữ liệu, kiểm tra NULL, tổng hợp theo nhóm.", "Viết query đúng, đọc bảng và tạo báo cáo tổng hợp cơ bản."),
    ("LEVEL 2", "JOIN & Relational Thinking", "Kết nối dữ liệu từ nhiều bảng và hiểu quan hệ", "INNER/LEFT/RIGHT/FULL/CROSS/SELF JOIN, PK/FK, 1-1, 1-N, N-N, duplicate do JOIN", "Ghép customer → orders → order_items → products; phát hiện JOIN làm phình số dòng.", "Tự thiết kế JOIN đúng grain và giải thích vì sao số dòng thay đổi."),
    ("LEVEL 3", "Subquery & CTE", "Chia query phức tạp thành các bước logic", "Subquery, EXISTS, NOT EXISTS, IN, ANY, ALL, WITH, multiple CTE, CTE chaining, recursive CTE", "Làm pipeline SQL nhiều bước: lọc → chuẩn hóa → aggregate → final dataset.", "Biến một query dài thành các bước dễ đọc, dễ debug và tái sử dụng trong transformation."),
    ("LEVEL 4", "Window Functions", "Phân tích theo dòng mà không làm mất từng record", "OVER, PARTITION BY, ORDER BY, ROW_NUMBER, RANK, DENSE_RANK, NTILE, LAG, LEAD, SUM/AVG/COUNT OVER, window frame", "Top N mỗi nhóm, dedup, running total, previous/next value, MoM/YoY.", "Tự giải bài toán ranking, lịch sử, cumulative và CTE + Window."),
    ("LEVEL 5", "Date/Time & Data Cleaning", "Xử lý thời gian và dữ liệu bẩn", "DATE, TIMESTAMP, TIMESTAMPTZ, CURRENT_DATE, INTERVAL, EXTRACT, DATE_TRUNC, AGE, TO_CHAR, timezone, NULL, duplicate, type conversion, JSON/ARRAY", "Doanh thu ngày/tuần/tháng, rolling 7 days, timezone, chuẩn hóa dữ liệu API.", "Tự xử lý các bài toán thời gian trong ETL/ELT và dữ liệu không sạch."),
    ("LEVEL 6", "Database Fundamentals", "Hiểu database chứ không chỉ viết query", "PK, FK, UNIQUE, NOT NULL, CHECK, DEFAULT, transaction, BEGIN/COMMIT/ROLLBACK, ACID, isolation, locks, deadlock", "Đảm bảo một batch load hoặc update nhiều bảng hoặc thành công toàn bộ hoặc rollback.", "Thiết kế constraint hợp lý và giải thích transaction/concurrency."),
    ("LEVEL 7", "Index & Query Optimization", "Làm query nhanh và đọc query plan", "B-tree, composite/partial index, GIN/GiST, EXPLAIN, EXPLAIN ANALYZE, Seq Scan, Index Scan, Bitmap Scan, Join strategies", "Tối ưu query ETL/report chạy chậm; tạo index đúng cột lọc/join; kiểm tra plan trước/sau.", "Tìm bottleneck và chứng minh query nhanh hơn bằng execution plan."),
    ("LEVEL 8", "Views, Functions & Procedures", "Đóng gói logic trong database", "VIEW, MATERIALIZED VIEW, FUNCTION, STORED PROCEDURE, TRIGGER, PL/pgSQL", "Tạo lớp dữ liệu dùng chung, refresh aggregate, đóng gói tác vụ database.", "Biết khi nào nên dùng view/function/procedure và khi nào không nên."),
    ("LEVEL 9", "ETL / ELT SQL", "Biến SQL thành công cụ xây data pipeline", "RAW/STAGING/CORE/MART, cleaning, validation, dedup, upsert, INSERT ON CONFLICT, incremental load, merge", "API/CSV → raw → staging → transform → fact/dim; chỉ xử lý dữ liệu mới bằng watermark.", "Thiết kế pipeline SQL có idempotency, validation và incremental processing."),
    ("LEVEL 10", "Data Warehouse", "Thiết kế dữ liệu phục vụ analytics", "OLTP vs OLAP, warehouse, mart, fact, dimension, grain, star/snowflake schema, natural/surrogate key, SCD Type 1/2", "Xây fact_orders và dim_customer; lưu lịch sử thay đổi customer bằng SCD2.", "Chọn grain đúng, thiết kế star schema và xử lý lịch sử dimension."),
    ("LEVEL 11", "Advanced Data Engineering SQL", "Giải bài toán dữ liệu phức tạp ở quy mô lớn", "Recursive CTE, advanced window frames, gaps & islands, sessionization, CDC, watermark, late-arriving data, partitioning, data quality, advanced optimization", "Sessionize event, incremental CDC, xử lý late data, partition bảng lớn và kiểm tra chất lượng dữ liệu.", "Tự thiết kế transformation/pipeline phức tạp và giải thích trade-off."),
]

wb = Workbook()
ws = wb.active
ws.title = "TỔNG QUAN"

# Theme
navy = "17324D"
blue = "1976D2"
cyan = "00A6A6"
green = "2E7D32"
orange = "EF6C00"
purple = "7B1FA2"
red = "C62828"
yellow = "F9A825"
light = "F4F7FB"
white = "FFFFFF"
dark = "263238"
gray = "607D8B"
thin = Side(style="thin", color="D9E2EC")

# Summary sheet
ws.merge_cells("A1:H2")
ws["A1"] = "SQL ROADMAP — DATA ENGINEER"
ws["A1"].font = Font(size=24, bold=True, color=white)
ws["A1"].fill = PatternFill("solid", fgColor=navy)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A4:H4")
ws["A4"] = "QUY TRÌNH HỌC"
ws["A4"].font = Font(size=14, bold=True, color=white)
ws["A4"].fill = PatternFill("solid", fgColor=blue)
ws["A4"].alignment = Alignment(horizontal="center")

flow = [
    ("1", "FOUNDATION", "Basic SQL"),
    ("2", "RELATIONS", "JOIN"),
    ("3", "TRANSFORM", "Subquery + CTE"),
    ("4", "ANALYTICS", "Window Functions"),
    ("5", "TIME", "Date/Time + Cleaning"),
    ("6", "DATABASE", "Transaction + Constraints"),
    ("7", "PERFORMANCE", "Index + EXPLAIN"),
    ("8", "PACKAGING", "View + Function + Procedure"),
    ("9", "PIPELINE", "ETL / ELT"),
    ("10", "WAREHOUSE", "Fact + Dimension + SCD"),
    ("11", "ADVANCED DE", "CDC + Incremental + Partition"),
]
for i, (n, title, sub) in enumerate(flow, start=6):
    ws.cell(i,1).value = n
    ws.cell(i,2).value = title
    ws.cell(i,3).value = "→"
    ws.cell(i,4).value = sub
    for c in range(1,5):
        ws.cell(i,c).border = Border(bottom=thin)
        ws.cell(i,c).alignment = Alignment(vertical="center")
    ws.cell(i,1).font = Font(size=12, bold=True, color=white)
    ws.cell(i,1).fill = PatternFill("solid", fgColor=[blue,cyan,green,purple,orange,red,yellow,blue,green,purple,navy][i-6])
    ws.cell(i,2).font = Font(bold=True, color=dark)
    ws.cell(i,3).font = Font(size=14, bold=True, color=gray)

ws.merge_cells("F6:H16")
ws["F6"] = (
    "TƯ DUY DATA ENGINEER\n\n"
    "Query dữ liệu\n↓\n"
    "Kết nối & biến đổi\n↓\n"
    "Phân tích theo dòng\n↓\n"
    "Xử lý thời gian & dữ liệu bẩn\n↓\n"
    "Đảm bảo tính đúng & transaction\n↓\n"
    "Tối ưu hiệu năng\n↓\n"
    "Xây ETL / ELT\n↓\n"
    "Thiết kế Data Warehouse\n↓\n"
    "Xử lý pipeline nâng cao"
)
ws["F6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["F6"].font = Font(size=12, bold=True, color=navy)
ws["F6"].fill = PatternFill("solid", fgColor="EAF4FF")
ws["F6"].border = Border(left=Side(style="medium", color=blue), right=Side(style="medium", color=blue),
                         top=Side(style="medium", color=blue), bottom=Side(style="medium", color=blue))

ws.merge_cells("A18:H18")
ws["A18"] = "CHUẨN ĐẦU RA TỔNG THỂ"
ws["A18"].font = Font(size=14, bold=True, color=white)
ws["A18"].fill = PatternFill("solid", fgColor=green)
ws["A18"].alignment = Alignment(horizontal="center")

outcomes = [
    "Viết SQL chắc từ cơ bản đến nâng cao.",
    "Hiểu grain, quan hệ bảng và tránh duplicate do JOIN.",
    "Dùng CTE + Window Function để xây transformation nhiều bước.",
    "Xử lý Date/Time, NULL, duplicate, JSON/ARRAY trong dữ liệu thực tế.",
    "Đọc EXPLAIN ANALYZE và tối ưu query.",
    "Thiết kế ETL/ELT, incremental load, upsert và data quality.",
    "Thiết kế Fact/Dimension, Star Schema và SCD.",
    "Giải được bài toán CDC, sessionization, late-arriving data và partitioning."
]
for r, text in enumerate(outcomes, start=20):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r,1).value = "✓ " + text
    ws.cell(r,1).font = Font(size=11, color=dark)
    ws.cell(r,1).alignment = Alignment(vertical="center")
    ws.cell(r,1).fill = PatternFill("solid", fgColor="F8FAFC")

# Level sheets
colors = [blue, cyan, green, purple, orange, red, yellow, blue, green, purple, navy]
for idx, (level, title, goal, topics, real, output) in enumerate(levels):
    sh = wb.create_sheet(f"{level.split()[0]} {idx+1}")
    accent = colors[idx]
    sh.merge_cells("A1:F2")
    sh["A1"] = f"{level}  |  {title}"
    sh["A1"].font = Font(size=20, bold=True, color=white)
    sh["A1"].fill = PatternFill("solid", fgColor=accent)
    sh["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    rows = [
        ("Mục tiêu", goal),
        ("Kiến thức cần học", topics),
        ("Ứng dụng thực tế", real),
        ("Chuẩn đầu ra", output),
    ]
    for r, (label, value) in enumerate(rows, start=4):
        sh.cell(r,1).value = label
        sh.cell(r,1).font = Font(bold=True, color=white)
        sh.cell(r,1).fill = PatternFill("solid", fgColor=accent)
        sh.cell(r,1).alignment = Alignment(vertical="top", wrap_text=True)
        sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        sh.cell(r,2).value = value
        sh.cell(r,2).alignment = Alignment(vertical="top", wrap_text=True)
        sh.cell(r,2).fill = PatternFill("solid", fgColor="F8FAFC")
        sh.cell(r,2).font = Font(color=dark)
    
    sh.merge_cells("A10:F10")
    sh["A10"] = "CHECKLIST HỌC TẬP"
    sh["A10"].font = Font(size=13, bold=True, color=white)
    sh["A10"].fill = PatternFill("solid", fgColor=accent)
    sh["A10"].alignment = Alignment(horizontal="center")
    
    checklist = [
        "Hiểu bản chất, không chỉ học cú pháp",
        "Tự viết query không nhìn đáp án",
        "Giải ít nhất 10 bài tập từ dễ → khó",
        "Làm 1 bài thực tế liên quan Data Engineering",
        "Tự giải thích query bằng lời",
        "Kiểm tra edge cases và dữ liệu NULL/duplicate"
    ]
    for r, item in enumerate(checklist, start=12):
        sh.cell(r,1).value = "☐"
        sh.cell(r,2).value = item
        sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        sh.cell(r,1).font = Font(size=13, color=accent)
        sh.cell(r,2).alignment = Alignment(wrap_text=True)
    
    # Practical scenario
    sh.merge_cells("A20:F20")
    sh["A20"] = "BÀI TOÁN THỰC TẾ GỢI Ý"
    sh["A20"].font = Font(size=13, bold=True, color=white)
    sh["A20"].fill = PatternFill("solid", fgColor=orange)
    sh["A20"].alignment = Alignment(horizontal="center")
    sh.merge_cells("A21:F23")
    scenario = {
        0: "Tạo báo cáo doanh thu theo phòng ban: lọc dữ liệu, xử lý NULL, GROUP BY và HAVING.",
        1: "Ghép customers → orders → order_items → products và kiểm tra vì sao JOIN tạo duplicate.",
        2: "Dùng CTE tách pipeline: raw orders → clean orders → customer revenue → final report.",
        3: "Tìm Top 3 khách hàng mỗi tháng, dedup bản ghi và tính doanh thu tháng trước bằng LAG().",
        4: "Tính doanh thu theo tháng, MoM, rolling 7 days; chuẩn hóa timestamp và timezone.",
        5: "Chạy một batch update nhiều bảng; nếu lỗi ở giữa phải ROLLBACK để dữ liệu không dở dang.",
        6: "Một query report chạy 15 giây: dùng EXPLAIN ANALYZE, tìm Seq Scan và thiết kế index phù hợp.",
        7: "Tạo VIEW cho BI và FUNCTION/PROCEDURE cho một tác vụ database lặp lại.",
        8: "Xây pipeline RAW → STAGING → CORE → MART, dedup và incremental load bằng watermark/upsert.",
        9: "Thiết kế fact_orders + dim_customer + dim_product; xử lý customer đổi thuộc tính bằng SCD2.",
        10: "Xử lý CDC/incremental events, sessionization, late-arriving data và partition bảng lớn."
    }[idx]
    sh["A21"] = scenario
    sh["A21"].alignment = Alignment(wrap_text=True, vertical="center")
    sh["A21"].fill = PatternFill("solid", fgColor="FFF8E1")
    
    sh.freeze_panes = "A4"
    widths = {"A":22,"B":22,"C":18,"D":18,"E":18,"F":18}
    for col, width in widths.items():
        sh.column_dimensions[col].width = width
    for row in range(1,24):
        sh.row_dimensions[row].height = 28
    sh.row_dimensions[1].height = 30
    sh.row_dimensions[2].height = 30
    sh.row_dimensions[5].height = 65
    sh.row_dimensions[6].height = 75
    sh.row_dimensions[7].height = 65
    sh.row_dimensions[21].height = 70
    for row in sh.iter_rows(min_row=4, max_row=23, min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(bottom=thin)

# General formatting
for sh in wb.worksheets:
    sh.sheet_view.showGridLines = False
    sh.freeze_panes = sh.freeze_panes or "A4"

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 3
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 22
ws.column_dimensions["H"].width = 22
for r in range(1, 29):
    ws.row_dimensions[r].height = 28
ws.row_dimensions[1].height = 34
ws.row_dimensions[2].height = 34
ws.sheet_view.showGridLines = False

wb.save(path)
print(path)
